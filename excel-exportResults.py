import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from optimization_engine import get_optimization_results
from constants import (BUFFER_MINUTES, ALPHA)
from time_utils import calculate_customer_satisfaction

def format_time(minutes): #format time
    hours = int(minutes // 60)
    mins = int(minutes % 60)
    return f"{hours:02d}:{mins:02d}"

def create_routes_comparison_sheet(wb, all_routes, all_arrival_times, days): #sheet showing five routes side by side
    ws = wb.create_sheet("Routes Comparison")
    ws.title = "Routes Comparison"

    for col in range(1, 16): ws.column_dimensions[get_column_letter(col)].width = 12 #column widths
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") #create headers
    header_font = Font(bold=True)
    ws.cell(row=1, column=1, value="Route Comparison").font = Font(bold=True, size=14)
    ws.cell(row=1, column=6,
            value=f"Alpha = {ALPHA} (Customer weight: {ALPHA}, Driver weight: {1.0 - ALPHA})").font = Font(bold=True)
    col = 1
    for day in days:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 2)
        ws.cell(row=2, column=col, value=day).font = header_font
        ws.cell(row=2, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col).fill = header_fill
        ws.cell(row=3, column=col, value="Stop #").font = header_font
        ws.cell(row=3, column=col).fill = header_fill
        ws.cell(row=3, column=col + 1, value="Customer ID").font = header_font
        ws.cell(row=3, column=col + 1).fill = header_fill
        ws.cell(row=3, column=col + 2, value="Arrival Time").font = header_font
        ws.cell(row=3, column=col + 2).fill = header_fill
        col += 3

    max_stops = max([len(route) for route in all_routes]) #data for each day
    for day_idx, day in enumerate(days):
        route = all_routes[day_idx]
        arrival_times = all_arrival_times[day_idx]
        col = day_idx * 3 + 1

        for stop_idx in range(len(route)):
            customer = route[stop_idx]
            row = stop_idx + 4
            ws.cell(row=row, column=col, value="D" if customer.id == 0 else stop_idx)
            ws.cell(row=row, column=col + 1, value=customer.id)
            if stop_idx < len(arrival_times):
                arrival_time = arrival_times[stop_idx]
                ws.cell(row=row, column=col + 2, value=format_time(arrival_time))

    summary_row = max_stops + 5 #summary
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)

    for day_idx, day in enumerate(days):
        col = day_idx * 3 + 1
        route = all_routes[day_idx]
        customers_served = len(route) - 2
        ws.cell(row=summary_row + 1, column=col, value="Customers:")
        ws.cell(row=summary_row + 1, column=col + 1, value=customers_served)
        if len(all_arrival_times[day_idx]) >= 2:
            start_time = all_arrival_times[day_idx][0]
            end_time = all_arrival_times[day_idx][-1]
            duration_hrs = (end_time - start_time) / 60
            ws.cell(row=summary_row + 2, column=col, value="Duration:")
            ws.cell(row=summary_row + 2, column=col + 1, value=f"{duration_hrs:.2f} hrs")
            ws.cell(row=summary_row + 3, column=col, value="Start:")
            ws.cell(row=summary_row + 3, column=col + 1, value=format_time(start_time))
            ws.cell(row=summary_row + 4, column=col, value="End:")
            ws.cell(row=summary_row + 4, column=col + 1, value=format_time(end_time))
    return ws

def create_customer_details_sheet(wb, all_routes, all_arrival_times, days): #customer service details
    ws = wb.create_sheet("Customer Service Details")
    ws.title = "Customer Service Details"

    columns = ["Customer ID", "Day", "Time Window", "Arrival Time", "Status", "Deviation", "Satisfaction"]
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") #create headers
    header_font = Font(bold=True)
    ws.cell(row=1, column=1, value="Customer Service Details").font = Font(bold=True, size=14)
    ws.cell(row=1, column=3,
            value=f"Alpha = {ALPHA} (Customer weight: {ALPHA}, Driver weight: {1.0 - ALPHA})").font = Font(bold=True)
    for col, header in enumerate(columns, 1):
        ws.cell(row=2, column=col, value=header).font = header_font
        ws.cell(row=2, column=col).fill = header_fill

    row = 3 #data
    for day_idx, day in enumerate(days):
        route = all_routes[day_idx]
        arrival_times = all_arrival_times[day_idx]
        for i in range(1, len(route) - 1):  #without depot
            customer = route[i]
            arrival_time = arrival_times[i]
            time_window = customer.time_windows[day_idx]
            satisfaction = calculate_customer_satisfaction(arrival_time, time_window, BUFFER_MINUTES)

            if arrival_time < time_window.start: #status and deviation
                deviation = time_window.start - arrival_time
                status = "Early"
            elif arrival_time > time_window.end:
                deviation = arrival_time - time_window.end
                status = "Late"
            else:
                deviation = 0
                status = "On Time"

            time_window_str = f"{format_time(time_window.start)}-{format_time(time_window.end)}" #format data
            deviation_str = f"{deviation // 60}h {deviation % 60}m" if deviation > 0 else "0m"

            if satisfaction >= 0.9:
                sat_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif satisfaction >= 0.7:
                sat_fill = PatternFill(start_color="C1FFC1", end_color="C1FFC1", fill_type="solid")
            elif satisfaction >= 0.5:
                sat_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif satisfaction >= 0.3:
                sat_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
            else:
                sat_fill = PatternFill(start_color="FF6961", end_color="FF6961", fill_type="solid")
            ws.cell(row=row, column=1, value=customer.id)
            ws.cell(row=row, column=2, value=day)
            ws.cell(row=row, column=3, value=time_window_str)
            ws.cell(row=row, column=4, value=format_time(arrival_time))
            ws.cell(row=row, column=5, value=status)
            ws.cell(row=row, column=6, value=deviation_str)
            ws.cell(row=row, column=7, value=f"{satisfaction:.2f}")
            ws.cell(row=row, column=7).fill = sat_fill
            row += 1
    return ws

def create_customers_by_day_sheet(wb, all_routes, all_arrival_times, all_unvisited_customers, days): #which customers are served on which days
    ws = wb.create_sheet("Customers By Day")
    ws.title = "Customers By Day"

    ws.column_dimensions['A'].width = 15 #column widths
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") #headers
    header_font = Font(bold=True)
    ws.cell(row=1, column=1, value="Customer Visits By Day").font = Font(bold=True, size=14)
    ws.cell(row=1, column=3,
            value=f"Alpha = {ALPHA} (Customer weight: {ALPHA}, Driver weight: {1.0 - ALPHA})").font = Font(bold=True)
    ws.cell(row=2, column=1, value="Customer ID").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    for col, day in enumerate(days, 2):
        ws.cell(row=2, column=col, value=day).font = header_font
        ws.cell(row=2, column=col).fill = header_fill

    all_customer_ids = set() #all unique customer IDs
    for day_idx in range(len(days)):
        route = all_routes[day_idx]
        for customer in route:
            if customer.id != 0:
                all_customer_ids.add(customer.id)
        for customer in all_unvisited_customers[day_idx]:
            all_customer_ids.add(customer.id)
    sorted_customer_ids = sorted(all_customer_ids)

    customer_rows = {} #customer ID rows
    for i, customer_id in enumerate(sorted_customer_ids):
        row = i + 3
        customer_rows[customer_id] = row
        ws.cell(row=row, column=1, value=customer_id)

    for day_idx, day in enumerate(days): #filling data for each day
        route = all_routes[day_idx]
        arrival_times = all_arrival_times[day_idx]
        col = day_idx + 2

        for i in range(1, len(route) - 1): #mark visited customers, skip depot
            customer = route[i]
            row = customer_rows[customer.id]
            arrival_time = arrival_times[i]
            satisfaction = calculate_customer_satisfaction(arrival_time, customer.time_windows[day_idx], BUFFER_MINUTES)

            if satisfaction >= 0.9: #colorcoading satisfaction
                fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif satisfaction >= 0.7:
                fill = PatternFill(start_color="C1FFC1", end_color="C1FFC1", fill_type="solid")
            elif satisfaction >= 0.5:
                fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif satisfaction >= 0.3:
                fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
            else:
                fill = PatternFill(start_color="FF6961", end_color="FF6961", fill_type="solid")
            ws.cell(row=row, column=col, value=format_time(arrival_time))
            ws.cell(row=row, column=col).fill = fill

    summary_row = len(sorted_customer_ids) + 4 #summary
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    for day_idx, day in enumerate(days):
        col = day_idx + 2
        route = all_routes[day_idx]
        visits = len(route) - 2
        total_customers = len(sorted_customer_ids)
        percent_served = (visits / total_customers) * 100 if total_customers > 0 else 0

        ws.cell(row=summary_row + 1, column=1, value="Customers Served:")
        ws.cell(row=summary_row + 1, column=col, value=visits)
        ws.cell(row=summary_row + 2, column=1, value="Total Customers:")
        ws.cell(row=summary_row + 2, column=col, value=total_customers)
        ws.cell(row=summary_row + 3, column=1, value="Percent Served:")
        ws.cell(row=summary_row + 3, column=col, value=f"{percent_served:.1f}%")
    return ws

def create_excel_report(): #create excel workbook
    print(f"Generating Excel report for alpha = {ALPHA}.")

    results = get_optimization_results()
    all_routes = results['routes']
    all_arrival_times = results['arrival_times']
    days = results['days']
    all_unvisited_customers = results['unvisited_customers']

    wb = Workbook() #create workbook
    wb.remove(wb.active)  #remove default sheet

    create_routes_comparison_sheet(wb, all_routes, all_arrival_times, days) #create sheets
    create_customer_details_sheet(wb, all_routes, all_arrival_times, days)
    create_customers_by_day_sheet(wb, all_routes, all_arrival_times, all_unvisited_customers, days)

    filename = f"route_optimization_report_alpha_{ALPHA}.xlsx" #save
    wb.save(filename)
    print(f"Excel report saved: {filename}")

if __name__ == "__main__":
    create_excel_report()